import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import numpy as np
import json
import os
import pandas as pd
from openpyxl.utils import get_column_letter


folder_path = "models/llama/llama_output"   

fixed_rows = [ 
                    "CURRENT ASSETS","Cash in the bank & hand", "Trade debtors", "Tax receivables", 
                    "Trade Debtors - Related party", "Provision for bad debts", 
                    "Stocks / Inventory", "Due from connected companies", 
                    "Due from Directors", "Deposits & Prepayments", 
                    "Leasehold Land & Builsding", "Retention Advances", 
                    "Available for sale investments", "Other Financial Assets", 
                    "Dividend Receivables of company", "Other receivables", 
                    "Short term Investments", "Other Current Assets", 
                    "TOTAL ASSETS","TRADE AND OTHER RECEIVABLES","CURRENT LIABILITIES","Due to banks- Overdrafts", "Due to banks-Short term borrowings", 
                    "Current Maturity on term loans", "Obligations under finance lease <1 Yr.", 
                    "Other Financial Liabilities", "Trade Creditors", "Trade Creditors - Related Party", 
                    "Due to connected companies", "Due to Directors", "Dividend / Tax payable", 
                    "Retentions", "Advances From Clients", "Mobilization Advances on Contracts", 
                    "Accrued Expenses", "Other Current Liabilities", "TOTAL CURRENT LIABILITIES","FIXED & TERM ASSETS",
                    "Freehold Land & Building","Leasehold Land & Building","Plant and, Machinery","Furniture, Fittings, Equipments",
                    "Motor vehicles","Operating Lease Assets","Capital work in progress","Investments (Incl.Debentures,Shares etc.)",
                    "Investment in connected companies","Lease/Hire Purchase receivables >1Yr","Other receivables > 1Yr",
                    "Intangible Assets (Goodwill)","Other fixed assets","FIXED & TERM ASSETS","NET ASSETS FINANCED BY:",
                    "Paid-up Capital","Preference Capital","Share Premium","General Reserves","Revaluation Reserves","Statutory Reserves Fund",
                    "Tax Reserves","Retained Profits","NET WORTH","LONG TERM LIABILITIES","Bank borrowings","Obligations under finance lease > 1 Yr.",
                    "Related party borrowings","Deposits maturing > 1Yr.","Debentures","Deferred tax","Retirement benefit obligations","Interest income",
                    "Interest expenses","Gross Profit","Other operating income","Taxes","Personnel costs","Other Administration expenses","Depreciation",
                    "Operating expenses","Direct expenses","Operating Profit","Non operating income / expenses","Provision for losses in advances",
                    "Profit / (loss) before tax","Share of results of associate","Tax","Net profit / (loss)","Dividends","Prior year adjustment (Revaluation)",
                    "Retained Earnings","Gross Non performing loans and advances","Provision for bad & doubtful debts","Net non performing loans & advances","Turnover",
        "Cost of sales (-)",
        "Gross profit",
        "Other operating income",
        "Distribution expenses (-)",
        "Administrative expenses (-)",
        "Litigation & compensation expenses (-)",
        "Other operating expenses (-)",
        "Operating Profit",
        "Non operating income",
        "Non operating expenses (-)",
        "Extra ordinary items - Income",
        "Extra ordinary items - Expenses (-)",
        "Finance Income",
        "Finance Cost - financial liabilities (-)",
        "Finance Cost - other (-)",
        "Profit / (loss) before tax",
        "Share of results of associate (+/-)",
        "Tax (+/-)",
        "Profit/(Loss) from continuing operations",
        "Net Profit/Loss from discontinued operations (+/-)",
        "Profit for the year",
        "Other net comprehensive income",
        "Dividends (-)",
        "Equity related income/ costs (+/-)",
        "Transfer to other Reserves (-)",
        "Prior year adjustment  - Income",
        "Prior year adjustment - Expense (-)",
        "Retained Earnings",
        "Interest expenses",
        "Depreciation"          
]


def fix_json_parent_name(data, filename):
  """
  Checks the first parent name and replaces with the JSON file name.

  Args:
    data: The JSON data loaded as a dictionary.
    filename: The name of the JSON file.

  Returns:
    The modified JSON data with the first parent name replaced by the filename.
  """ 
  first_parent_name = list(data.keys())[0]

  if first_parent_name not in fixed_rows:   
    filename_parts = filename.split("_")    
    file_name =  filename_parts[0]
    data[file_name] = data.pop(first_parent_name)

  return data

def process_json_folder(folder_path):
    """
    Processes all JSON files in a folder and replaces the first parent name with the filename.

    Args:
        folder_path: The path to the folder containing JSON files.

    Returns:
        List of processed JSON filenames.
    """
def process_json_folder(folder_path):
    """
    Processes all JSON files in a folder and replaces the first parent name with the filename.

    Args:
        folder_path: The path to the folder containing JSON files.

    Returns:
        List of processed JSON filenames.
    """
    processed_files = []  
    for filename in os.listdir(folder_path):
        if filename.endswith(".json"):
            file_path = os.path.join(folder_path, filename)
            try:
                with open(file_path, "r") as json_file:
                    data = json.load(json_file)
                    modified_data = fix_json_parent_name(data.copy(), filename)
    
                with open(file_path, "w") as json_file:
                    json.dump(modified_data, json_file, indent=2)  
                print(f"Processed file: {filename}")
                processed_files.append(filename)  
            except FileNotFoundError:
                print(f"Error: File not found - {file_path}")
    return processed_files  

processed_files = process_json_folder(folder_path)

def add_json_data(existing_data, new_data, row_mapping=None, column_mapping=None):
    """
    Add new JSON data to existing JSON data. If keys are the same, perform addition.

    Parameters:
        existing_data (dict): Existing JSON data.
        new_data (dict): New JSON data to be added.
        row_mapping (dict): Mapping dictionary for JSON row keys to fixed row names.
        column_mapping (dict): Mapping dictionary for JSON column keys to fixed column names.

    Returns:
        dict: Combined JSON data with additions for same keys.
    """
    for key, value in new_data.items():
        try:
            if key in existing_data:
                if isinstance(value, dict):  
                    for sub_key, sub_value in value.items():
                        if isinstance(sub_value, dict):
                            existing_data[key][sub_key] = add_json_data(existing_data[key].get(sub_key, {}), sub_value, row_mapping, column_mapping)
                        else:
                            try:
                                sub_value = int(sub_value)
                            except ValueError:
                                raise ValueError("Non-integer value encountered")
                            existing_data[key][sub_key] = existing_data[key].get(sub_key, 0) + sub_value
                else:
                    existing_data[key] += value  
            else:
                existing_data[key] = value
        except ValueError:
            return existing_data

    if column_mapping:
        for row_key, row_data in existing_data.items():
            if isinstance(row_data, dict):
                for column_key, column_value in row_data.items():
                    if column_key in column_mapping:
                        existing_data[row_key][column_mapping[column_key]] = existing_data[row_key].pop(column_key)

    if row_mapping:
        existing_data_mapped = {}
        for row_key, row_data in existing_data.items():
            if row_key in row_mapping:
                existing_data_mapped[row_mapping[row_key]] = row_data
            else:
                existing_data_mapped[row_key] = row_data
        existing_data = existing_data_mapped

    return existing_data

json_files = [os.path.join(folder_path, file) for file in os.listdir(folder_path) if file.endswith(".json")]

combined_data_bs = {}  
combined_data_notes = {}  

key_to_row_mapping = {
    "Dividend Receivables":  "Dividend Receivables of company",
    "Export":  "Dividend Receivables of company",
    "Cash in bank & hand": "Cash in the bank & hand",
    "Deferred taxation": "Deferred tax",
    "Depreciation": "Leasehold Land & Building",
    "Income statement" : "Trade debtors",
    "Revenue" : "Tax receivables",
    "DEFFERED TAX" : "Deferred tax",
    "BANK OVERDRAFTS": "Due to banks- Overdrafts",
    "Lease Interest": "Cash and cash equivalents",
    "Overdraft Interest": "Due to banks- Overdrafts",
    "Total Inventories": "Stocks / Inventory",
    "RETIREMENT BENEFIT OBLIGATION": "Other receivables",
    "On Investments": "Short term Investments",
    "Sundry Income": "Interest income",
    "Property Plant And Equipment": "Other Financial Assets",

    
}

key_to_column_mapping = {
    "2020": "2020",
    "2021": "2021",
    "2022": "2022",
    "2023": "2023"
    
}

for json_file in json_files:
    with open(json_file, "r") as f:
        data = json.load(f)
    combined_data_bs = add_json_data(combined_data_bs, data)

fixed_rows = {
    
    "Income Statement": [
        "Turnover",
        "Cost of sales (-)",
        "Gross profit",
        " ",
        "Other operating income",
        "Distribution expenses (-)",
        "Administrative expenses (-)",
        "Litigation & compensation expenses (-)",
        "Other operating expenses (-)",
        "",
        "Operating Profit",
        "",
        "Non operating income",
        "Non operating expenses (-)",
        "Extra ordinary items - Income",
        "Extra ordinary items - Expenses (-)",
        "Finance Income",
        "Finance Cost - financial liabilities (-)",
        "Finance Cost - other (-)",
        "",
        "Profit / (loss) before tax",
        "Share of results of associate (+/-)",
        "Tax (+/-)",
        "",
        "Profit/(Loss) from continuing operations",
        "Net Profit/Loss from discontinued operations (+/-)",
        "",
        "Profit for the year",
        "Other net comprehensive income",
        "Dividends (-)",
        "Equity related income/ costs (+/-)",
        "Transfer to other Reserves (-)",
        "Prior year adjustment  - Income",
        "Prior year adjustment - Expense (-)",
        "",
        "Retained Earnings",
        "",
        "Interest expenses",
        "Depreciation",
            "",
        "% Income Statement",
        "Sales growth",
        "Cost of sales",
        "Gross profit",
        "Other operating income",
        "Distribution expenses",
        "Administrative expenses",
        "Litigation & compensation expenses",
        "Other operating expenses",
        "Operating Profit",
        "Non operating income",
        "Non operating expenses",
        "Finance Cost - financial liabilities",
        "Profit / (loss) before tax",
        "Tax",
        "Net profit / (loss)",
        "",
        "",
        ""
        "CASHFLOW STATEMENT",
        "Net profit / (loss) before taxation",
        "Adjustments for non cash transactions",
        "Operating profit before working capital changes",
        "Adjustment for working capital changes",
        "Cash generated from operations",
        "Net Interest paid",
        "Other payments (Gratuity / Tax etc.)",
        "Net cashflow from operating activities",
        "Net cashflow from investing activities",
        "Lease repayment",
        "Dividend payment",
        "Financing surplus / (Deficit)",
        "",
        "Net external financing     - Debt",
        "                           - Equity",
        "Net cash inflow / (outflow)",
        ""

    ]
}

notes_data_dict = {}
print(notes_data_dict)
 
for json_file in json_files:
    with open(json_file, "r") as f:
        data = json.load(f)
        first_parent_name = next(iter(data.keys()))
        for section, rows in fixed_rows.items():
            if first_parent_name in rows:
                if first_parent_name not in notes_data_dict:
                    notes_data_dict[first_parent_name] = {} 
                notes_data_dict[first_parent_name].update(data[first_parent_name])
                break  
        else:
            combined_data_bs 

combined_data_notes = {}
df_notes = pd.DataFrame()

for parent_name, data_dict in notes_data_dict.items():
    combined_data_notes[parent_name] = data_dict

if combined_data_notes:
    df_notes = pd.DataFrame(combined_data_notes)
    if len(processed_files) == len(df_notes):
        df_notes['File Name'] = processed_files
        df_notes['Page Number'] = [filename.split('_')[-1].split('.')[0] for filename in processed_files]
    else:
        print("Error: Length of processed_files does not match the length of df_notes index.")

def map_row_name(row_name):
    return key_to_row_mapping.get(row_name, row_name)

if not df_notes.empty:
    df_notes.index = df_notes.index.map(map_row_name)
    df_notes = df_notes.apply(pd.to_numeric, errors='coerce')

file_name_base = os.path.splitext(os.path.basename(json_files[0]))[0]
first_parent_name = next(iter(data.keys()))
sheet_name = next(iter(combined_data_bs))
sheet_name_notes = "Notes"
df_bs = pd.DataFrame(combined_data_bs[sheet_name]).transpose()

print(df_bs)
if df_bs is not None:
    df_bs.index = df_bs.index.map(map_row_name)
    df_bs = df_bs.apply(pd.to_numeric, errors='coerce')

sheet_fixed_rows = fixed_rows.keys()

year_columns = ['2020', '2021', '2022', '2023']

for year in year_columns:
    column_name = key_to_column_mapping.get(year, year)
    if column_name not in df_bs.columns:
        df_bs[column_name] = 0
        
df_bs = df_bs[[col for col in df_bs.columns if col in year_columns]]
df_bs = df_bs[year_columns]
def map_row_name(row_name):
    return key_to_row_mapping.get(row_name, row_name)
df_bs = df_bs.groupby(level=0).sum()
df = df_bs.reindex(sum(fixed_rows.values(), []))

section_totals = {}  
for section, rows in fixed_rows.items():
    valid_rows = [row for row in rows if row in df.index]
    section_df = df.loc[valid_rows]
    section_totals[section] = section_df.sum(axis=0)

for section, total_values in section_totals.items():
    total_row_name = fixed_rows[section][-2]  
    df.loc[total_row_name] = total_values.values  

def format_currency(value):
    if np.isnan(value):
        return ''
    return '{:,.2f}'.format(value)

df = df.applymap(format_currency)

for parent_name, json_data in combined_data_notes.items():
    total_column = {}
    for key, sub_dict in json_data.items():
        if key != 'Total':  
            for sub_key, value in sub_dict.items():
                if isinstance(value, dict):
                    print(f"Skipping key '{sub_key}' in parent '{parent_name}' because it is a dictionary.")
                    continue 
                try:
                    value = int(value)
                except (ValueError, TypeError):
                    print(f"Non-integer value '{value}' encountered for key '{sub_key}' in parent '{parent_name}'. Skipping addition.")
                    continue  
                total_column[sub_key] = total_column.get(sub_key, 0) + value
    json_data["Total"] = total_column

parent_table_column_totals = {}

for parent_name, json_data in combined_data_notes.items():
    total_column = {}

    for key, sub_dict in json_data.items():
        if key != 'Total': 
            for sub_key, value in sub_dict.items():
                if isinstance(value, dict):
                    print(f"Skipping key '{sub_key}' in parent '{parent_name}' because it is a dictionary.")
                    continue 
                try:
                    value = int(value)
                except (ValueError, TypeError):
                    print(f"Non-integer value '{value}' encountered for key '{sub_key}' in parent '{parent_name}'. Skipping addition.")
                    continue  
                total_column[sub_key] = total_column.get(sub_key, 0) + value
    parent_table_column_totals[parent_name] = total_column

for parent_table, column_totals in parent_table_column_totals.items():
    for column_name, total_value in column_totals.items():
        if column_name in df.columns:
            df.at[parent_table, column_name] = total_value  

for parent_table, column_totals in parent_table_column_totals.items():
    row_index = df.index.get_loc(parent_table)
    
    for column_name, total_value in column_totals.items():
        if column_name in df.columns:
            df.at[parent_table, column_name] = total_value

excel_file = "bsr/excel_files/income_statement.xlsx"

with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    if combined_data_bs:  
        df.to_excel(writer, sheet_name=sheet_name, index_label="INCOME STATEMENT", startrow=2)

    if combined_data_notes:  
        current_row_notes = 2  
        for parent_name, json_data in combined_data_notes.items():
            df_parent_notes = pd.DataFrame(json_data).transpose()  
            df_parent_notes.index.name = parent_name 
            df_parent_notes.to_excel(writer, sheet_name=sheet_name_notes, startrow=current_row_notes)
            current_row_notes += len(df_parent_notes) + 2  
    else: 
        df_empty_notes = pd.DataFrame(columns=["Empty Notes Section"])
        df_empty_notes.to_excel(writer, sheet_name=sheet_name_notes, index=False)

wb = load_workbook(excel_file)
ws3 = wb[sheet_name]
ws4 =wb[sheet_name_notes]

max_row = ws3.max_row
max_column = ws3.max_column
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

max_row = ws4.max_row
max_column = ws4.max_column
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
    for cell in row:
        cell.border = thin_border

for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=ws4.max_column):
    for cell in row:
        cell.border = thin_border

topic_cell1 = ws3.cell(row=1, column=1)
topic_cell1_value = file_name_base.split('_')[0]
topic_cell1.value= topic_cell1_value
topic_cell1.alignment = Alignment(horizontal='center')
max_column_letter = get_column_letter(ws3.max_column)
topic_range = f'A1:{max_column_letter}1'
ws3.merge_cells(topic_range)

topic_cell2 = ws4.cell(row=1, column=1)
topic_cell3 = ws4.cell(row=2, column=3)
topic_cell2_value = file_name_base.split('_')[0]
topic_cell2.value= topic_cell2_value
topic_cell2.alignment = Alignment(horizontal='center')

for row in ws3.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=max_column):
    for cell in row:
        if cell.value in combined_data_notes:
            totals = combined_data_notes[cell.value]["Total"]
            for col_idx, col in enumerate(row):
                if col_idx > 0:  
                    col.value = totals.get(col.value, 0)
            break 

for row in ws3.iter_rows(min_row=3, max_row=500, min_col=1, max_col=max_column):
    for cell in row:
        for section, rows in fixed_rows.items():
            if cell.value in rows:
                cell.alignment = Alignment(horizontal='left')
                break 

for row in ws4.iter_rows(min_row=3, max_row=500 , min_col=1, max_col=max_column):
    for cell in row:
        for section, rows in fixed_rows.items():
            if cell.value in rows:
                cell.alignment = Alignment(horizontal='left')
                break  

for column_cells in ws3.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws3.column_dimensions[column_cells[1].column_letter].width = length + 2  

for column_cells in ws3.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws4.column_dimensions[column_cells[1].column_letter].width = length + 2  
