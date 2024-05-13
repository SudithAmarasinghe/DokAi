import concurrent.futures
import subprocess
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
from balance_sheet import sheet_name

class FinalBSR:
    def __init__(self, file1, file2):
        self.wb1 = load_workbook(file1)
        self.wb2 = load_workbook(file2)
        self.scripts = ["bsr/balance_sheet.py", "bsr/income_statement.py"]


    def run_script(self, script_name):
        subprocess.run(["python", script_name])

    def copy_data_between_sheets(self):
        for sheet_name in self.wb2.sheetnames:
            ws1 = self.wb1[sheet_name] if sheet_name in self.wb1.sheetnames else self.wb1.create_sheet(sheet_name)
            ws2 = self.wb2[sheet_name]

            start_col_ws2 = ws1.max_column + 2 if ws1.max_column > 0 else 1

            for row in range(1, ws2.max_row + 1):
                for col in range(1, ws2.max_column + 1):
                    ws1.cell(row=row, column=col + start_col_ws2 - 1, value=ws2.cell(row=row, column=col).value)

    def autofit_columns(self):
        for ws in self.wb1:
            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in ws.iter_rows(min_row=1, min_col=col, max_row=ws.max_row, max_col=col):
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

    def apply_styles(self):
        header_font = Font(bold=True)
        alignment = Alignment(horizontal='left')
        alignment_center = Alignment(horizontal='center')
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))

        for ws in self.wb1:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.font = header_font

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = alignment

            for cell in ws[3]:
                cell.alignment = alignment_center
                cell.font = Font(bold=True)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border_style

            if ws.max_row > 0 and ws.max_column > 0:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
                merged_cell = ws['A1']
                merged_cell.value =  sheet_name
                merged_cell.alignment = Alignment(horizontal='center')

            for cell in ws.iter_cols(min_col=6, max_col=6):
                for c in cell:
                    c.font = Font(bold=True)

    def calculate_sum_pairs(self, sum_pairs):
        for col, (cell1, cell2, dest_cell) in sum_pairs.items():
            for ws in self.wb1:
                cell1_value = ws[cell1].value if ws[cell1].value else 0
                cell2_value = ws[cell2].value if ws[cell2].value else 0
                ws[dest_cell] = cell1_value + cell2_value

    def calculate_additional_sums(self, additional_sum, sheet_name):
        ws = self.wb1[sheet_name]
        for dest_cell, expression in additional_sum.items():
            if isinstance(expression, tuple):  
                cell_range = ws[expression[0]] 
                sum_value = 0
                for row in cell_range:
                    for cell in row:
                        if cell.value:
                            if isinstance(cell.value, (int, float)):
                                sum_value += cell.value
                            elif isinstance(cell.value, str) and cell.value.isdigit():
                                sum_value += int(cell.value)
                            else:
                                print(f"Warning: Non-numeric value found in cell {cell.coordinate}. Skipping.")
                ws[dest_cell] = sum_value
            elif isinstance(expression, str) and expression.startswith('-'):  
                neg_cell_reference = expression[1:]  
                neg_cell_value = ws[neg_cell_reference].value if ws[neg_cell_reference].value else 0
                ws[dest_cell] = -neg_cell_value
            elif expression.startswith('='): 
                ws[dest_cell] = expression
            else: 
                try:
                    cell_value = eval(expression) 
                    ws[dest_cell] = cell_value
                except Exception as e:
                    print(f"Error evaluating expression for cell {dest_cell}: {e}")

    def save_and_run_scripts(self):
        self.wb1.save('download/BSR_SHEET.xlsx')
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(self.run_script, self.scripts)

    def run(self, file1, file2):
            self.wb1 = load_workbook(file1)
            self.wb2 = load_workbook(file2)

            self.copy_data_between_sheets()
            self.autofit_columns()
            self.apply_styles()

            additional_sum = {

                'H6': '=H5+H4',
                'I6': '=I5+I4',
                'J6': '=J5+J4',
                'K6': '=K5+K4', 

                'H14': ('H6:H12',),
                'I14': ('I6:I12',),
                'J14': ('J6:J12',),
                'K14': ('K6:K12',),

                'H24': ('H14:H22',),
                'I24': ('I14:I22',),
                'J24': ('J14:J22',),
                'K24': ('K14:K22',),

                'H28': ('H24:H27',),
                'I28': ('I24:I27',),
                'J28': ('J24:J27',),
                'K28': ('K24:K27',),

                'H31': ('H28:H30',),
                'I31': ('I28:I30',),
                'J31': ('J28:J30',),
                'K31': ('K28:K30',),

                'H39': ('H31:H38',),
                'I39': ('I31:I38',),
                'J39': ('J31:J38',),
                'K39': ('K31:K38',),

                'H42': '-H21',
                'I42': '-I21',
                'J42': '-J21',
                'K42': '-K21',

                'H46': '=-H5/H4',
                'I46': '=-I5/I4',
                'J46': '=-J5/J4',
                'K46': '=-K5/K4',          


                'H47': '=H6/H4',
                'I47': '=I6/I4',
                'J47': '=J6/J4',
                'K47': '=K6/K4',         

                'H48': '=H8/H4',
                'I48': '=I8/I4',
                'J48': '=J8/J4',
                'K48': '=K8/K4',

                'H49': '=-H9/H4',
                'I49': '=-I9/I4',
                'J49': '=-J9/J4',
                'K49': '=-K9/K4',   

                'H50': '=-H10/H4',
                'I50': '=-I10/I4',
                'J50': '=-J10/J4',
                'K50': '=-K10/K4',   

                'H50': '=-H11/H4',
                'I50': '=-I11/I4',
                'J50': '=-J11/J4',
                'K50': '=-K11/K4', 

                'H51': '=-H11/H4',
                'I51': '=-I11/I4',
                'J51': '=-J11/J4',
                'K51': '=-K11/K4',  

                'H52': '=-H12/H4',
                'I52': '=-I12/I4',
                'J52': '=-J12/J4',
                'K52': '=-K12/K4',  

                'H53': '=-H14/H4',
                'I53': '=-I14/I4',
                'J53': '=-J14/J4',
                'K53': '=-K14/K4',  

                'H54': '=H16/H4',
                'I54': '=I16/I4',
                'J54': '=J16/J4',
                'K54': '=K16/K4',      

                'H55': '=-H17/H4',
                'I55': '=-I17/I4',
                'J55': '=-J17/J4',
                'K55': '=-K17/K4', 

                'H56': '=-H21/H4',
                'I56': '=-I21/I4',
                'J56': '=-J21/J4',
                'K56': '=-K21/K4', 

                'H57': '=H24/H4',
                'I57': '=I24/I4',
                'J57': '=J24/J4',
                'K57': '=K24/K4',   

                'H58': '=-H26/H4',
                'I58': '=-I26/I4',
                'J58': '=-J26/J4',
                'K58': '=-K26/K4', 

                'H59': '=H32/H4',
                'I59': '=I32/I4',
                'J59': '=J32/J4',
                'K59': '=K32/K4',

                'H63': '=H24',
                'I63': '=I24',
                'J63': '=J24',
                'K63': '=K24',

                'H65': '=H63+H64' ,
                'I65': '=I63+I64',
                'J65': '=J63+J64',
                'K65': '=K63+K64',

                'H67': '=H65+H66' ,
                'I67': '=I65+I66',
                'J67': '=J65+J66',
                'K67': '=K65+K66',

                'H70': '=H67+H68+H69' ,
                'I70': '=I67+I68+I69',
                'J70': '=J67+J68+J69',
                'K70': '=K67+K68+K69',

                'H74': '=H70+H71+H72+H73' ,
                'I74': '=I70+I71+H72+I73',
                'J74': '=J70+J71+H72+J73',
                'K74': '=K70+K71+H72+K73',

                'H78': '=H75+H76+H77' ,
                'I78': '=I75+I76+I77',
                'J78': '=J75+J76+J77',
                'K78': '=K75+K76+K77'
                
            }

            
            self.calculate_additional_sums(additional_sum, sheet_name=sheet_name)
            self.save_and_run_scripts()
            